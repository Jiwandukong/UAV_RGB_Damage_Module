"""OBJ-ray quantitative exports and the public Excel deliverable."""

from __future__ import annotations

import csv
import json
from collections.abc import Mapping, Sequence
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import numpy as np
from PIL import Image

from .config import MODEL_VARIANT, CrackSegConfig
from .instances import summarize_class_mask_instances
from .metrics import validate_class_masks
from .visualization import (
    class_masks_to_channel_image,
    class_masks_to_color,
    make_multilabel_overlay,
)


DAMAGE_NAMES_KO = {"CRC": "균열", "DLM": "박리", "SPL": "박락"}

PRIMARY_COLUMNS = [
    "image",
    "damage_id",
    "damage_type",
    "damage_name_ko",
    "pixel_nodes_json",
    "world_center_x_m",
    "world_center_y_m",
    "world_center_z_m",
    "length_px",
    "length_m",
    "width_px",
    "width_m",
    "area_m2",
]

# The provided output example contains exactly these 13 columns.  Do not add
# diagnostics or QA fields to the user-facing CSV/XLSX without prior approval.
DETAIL_COLUMNS = PRIMARY_COLUMNS


def prediction_masks_by_id(
    prediction: Mapping[str, Any], config: CrackSegConfig
) -> dict[int, np.ndarray]:
    raw_masks = prediction.get("masks")
    if not isinstance(raw_masks, Mapping):
        raise ValueError("prediction must contain a masks mapping")
    result: dict[int, np.ndarray] = {}
    for class_name, class_id in config.class_ids.items():
        raw = raw_masks.get(class_name, raw_masks.get(class_id))
        if raw is None:
            raise ValueError(f"prediction is missing the {class_name} mask")
        mask = np.asarray(raw)
        if mask.dtype != np.bool_ or mask.ndim != 2:
            raise TypeError(f"{class_name} mask must be a two-dimensional bool array")
        result[int(class_id)] = mask
    validate_class_masks(result)
    return result


def save_image_result(
    *,
    image_rgb: np.ndarray,
    image_path: str | Path,
    prediction: Mapping[str, Any],
    config: CrackSegConfig,
    geo3d_context: Any,
    output_dir: str | Path,
    min_instance_area_px: int = 8,
    alpha: float | None = None,
    damage_id_start: int = 1,
) -> dict[str, Any]:
    """Save masks, overlay, quantitative table and provenance for one image."""
    source = Path(image_path)
    output_root = Path(output_dir)
    stem = source.stem
    class_masks = prediction_masks_by_id(prediction, config)
    height, width = validate_class_masks(class_masks)
    image_rgb = np.asarray(image_rgb, dtype=np.uint8)
    if image_rgb.shape != (height, width, 3):
        raise ValueError(
            "source image and prediction masks must share the untouched pixel grid: "
            f"image={image_rgb.shape}, masks={(height, width)}"
        )
    if min_instance_area_px < 1:
        raise ValueError("min_instance_area_px must be >= 1")
    if damage_id_start < 1:
        raise ValueError("damage_id_start must be >= 1")

    masks_dir = output_root / "masks" / stem
    overlays_dir = output_root / "overlays"
    tables_dir = output_root / "tables" / "per_image"
    metadata_dir = output_root / "metadata" / "per_image"
    for directory in (masks_dir, overlays_dir, tables_dir, metadata_dir):
        directory.mkdir(parents=True, exist_ok=True)

    combined_mask_path = masks_dir / "multilabel_rgb.png"
    preview_path = masks_dir / "preview.png"
    overlay_path = overlays_dir / f"{stem}_overlay.png"
    Image.fromarray(class_masks_to_channel_image(class_masks), mode="RGB").save(
        combined_mask_path
    )
    Image.fromarray(class_masks_to_color(class_masks, config.palette), mode="RGB").save(
        preview_path
    )
    for class_name, class_id in config.class_ids.items():
        binary_path = masks_dir / f"{class_name}.png"
        Image.fromarray(class_masks[class_id].astype(np.uint8) * 255, mode="L").save(
            binary_path
        )
    overlay_alpha = config.overlay_alpha if alpha is None else float(alpha)
    Image.fromarray(
        make_multilabel_overlay(
            image_rgb, class_masks, config.palette, alpha=overlay_alpha
        ),
        mode="RGB",
    ).save(overlay_path)

    raw_rows, raw_summary, label_maps = summarize_class_mask_instances(
        class_masks,
        config.class_names,
        min_area_px=min_instance_area_px,
    )
    spatial_rows = geo3d_context.image_regions_to_world3d(
        raw_rows,
        class_masks=class_masks,
        instance_label_maps=label_maps,
    )
    if len(spatial_rows) != len(raw_rows):
        raise RuntimeError("mesh ray mapper returned a different number of instances")

    details: list[dict[str, Any]] = []
    for local_number, (base, spatial) in enumerate(
        zip(raw_rows, spatial_rows), start=1
    ):
        merged = dict(base)
        merged.update(spatial)
        class_name = str(merged["class_name"])
        details.append(
            {
                "image": source.name,
                "damage_id": f"D{damage_id_start + local_number - 1:06d}",
                "damage_type": class_name,
                "damage_name_ko": DAMAGE_NAMES_KO.get(class_name, class_name),
                "pixel_nodes_json": merged.get("nodes_image_xy_json"),
                "world_center_x_m": merged.get("world_x_m"),
                "world_center_y_m": merged.get("world_y_m"),
                "world_center_z_m": merged.get("world_z_m"),
                "length_px": merged.get("length_px"),
                "length_m": merged.get("length_m"),
                "width_px": merged.get("width_px"),
                "width_m": merged.get("width_m"),
                "area_m2": merged.get("area_m2"),
            }
        )

    details_csv = tables_dir / f"{stem}_damage_details.csv"
    write_csv(details_csv, details, DETAIL_COLUMNS)
    class_by_id = {row["class_id"]: dict(row) for row in raw_summary}
    image_summary: dict[str, Any] = {
        "image": source.name,
        "width_px": width,
        "height_px": height,
        "damage_count": len(details),
        "mapped_damage_count": sum(
            row.get("xyz_valid") is True for row in spatial_rows
        ),
        "quantified_damage_count": sum(
            row.get("measurement_valid") is True for row in spatial_rows
        ),
    }
    for class_name, class_id in config.class_ids.items():
        class_row = class_by_id.get(class_id, {})
        image_summary[f"{class_name}_count"] = int(
            class_row.get("instance_count", 0)
        )
        image_summary[f"{class_name}_area_px"] = int(
            class_row.get("total_area_px", 0)
        )
    image_summary["CRC_total_length_m"] = _sum_values(details, "CRC", "length_m")
    image_summary["DLM_total_area_m2"] = _sum_values(details, "DLM", "area_m2")
    image_summary["SPL_total_area_m2"] = _sum_values(details, "SPL", "area_m2")

    metadata = {
        "schema_version": "1.0",
        "created_utc": datetime.now(timezone.utc).isoformat(),
        "source_image": source.name,
        "source_shape_hw": [height, width],
        "model_variant": MODEL_VARIANT,
        "checkpoint": json_ready(prediction.get("checkpoint")),
        "inference": json_ready(prediction.get("inference")),
        "class_prediction": json_ready(prediction.get("classes")),
        "mapping": json_ready(geo3d_context.to_dict()),
        "quantification": {
            "CRC": "minimum-area rotated rectangle extents times local directional mesh GSD",
            "DLM": "mask pixels times local mesh surface area per pixel",
            "SPL": "mask pixels times local mesh surface area per pixel",
            "coordinate_crs": "EPSG:5186 for X/Y; Z-up metres",
            "approximate": True,
            "minimum_component_area_px": min_instance_area_px,
        },
        "image_summary": image_summary,
        "warnings": [
            "The OBJ CRS is an external project contract; OBJ does not embed CRS metadata.",
            "The MRK Ellh-to-OBJ vertical datum relationship is not verified.",
            "DJI DewarpData is not applied by the current calibrated-pinhole ray model.",
            "Metrics are model predictions and require engineering review.",
        ],
    }
    metadata_path = metadata_dir / f"{stem}_metadata.json"
    metadata_path.write_text(
        json.dumps(metadata, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )
    return {
        "rows": details,
        "summary": image_summary,
        "paths": {
            "details_csv": details_csv,
            "overlay": overlay_path,
            "multilabel_mask": combined_mask_path,
            "metadata": metadata_path,
        },
    }


def write_combined_outputs(
    *,
    rows: list[dict[str, Any]],
    run_metadata: Mapping[str, Any],
    output_dir: str | Path,
) -> dict[str, Path]:
    output_root = Path(output_dir)
    tables = output_root / "tables"
    tables.mkdir(parents=True, exist_ok=True)
    for index, row in enumerate(rows, start=1):
        row["damage_id"] = f"D{index:06d}"
    details_csv = tables / "damage_results.csv"
    workbook_path = tables / "damage_results.xlsx"
    write_csv(details_csv, rows, DETAIL_COLUMNS)
    write_workbook(
        workbook_path,
        rows=rows,
    )
    run_metadata_path = output_root / "metadata" / "run_metadata.json"
    run_metadata_path.parent.mkdir(parents=True, exist_ok=True)
    run_metadata_path.write_text(
        json.dumps(json_ready(run_metadata), ensure_ascii=False, indent=2, sort_keys=True)
        + "\n",
        encoding="utf-8",
    )
    return {
        "damage_csv": details_csv,
        "excel": workbook_path,
        "run_metadata": run_metadata_path,
    }


def write_workbook(
    path: str | Path,
    *,
    rows: list[dict[str, Any]],
) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as exc:
        raise ImportError("openpyxl is required to write the Excel deliverable") from exc

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    detail_sheet = workbook.active
    detail_sheet.title = "Damage_Details"
    _append_table(detail_sheet, rows, DETAIL_COLUMNS)

    header_fill = PatternFill("solid", fgColor="5595AF")
    header_font = Font(color="FFFFFF", bold=True)
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for column_cells in sheet.columns:
            letter = column_cells[0].column_letter
            values = [str(cell.value or "") for cell in column_cells[:100]]
            width = min(max(max(map(len, values), default=8) + 2, 10), 60)
            sheet.column_dimensions[letter].width = width
        if sheet.title == "Damage_Details":
            for row in sheet.iter_rows(min_row=2):
                row[4].alignment = Alignment(wrap_text=False)
    workbook.save(path)


def _append_table(sheet: Any, rows: list[dict[str, Any]], columns: Sequence[str]) -> None:
    sheet.append(list(columns))
    for row in rows:
        sheet.append([_excel_value(row.get(column)) for column in columns])


def _excel_value(value: Any) -> Any:
    if isinstance(value, (dict, list, tuple)):
        return json.dumps(json_ready(value), ensure_ascii=False, separators=(",", ":"))
    if isinstance(value, (np.bool_, bool)):
        return bool(value)
    if isinstance(value, np.integer):
        return int(value)
    if isinstance(value, np.floating):
        return float(value)
    return value


def _sum_values(rows: list[dict[str, Any]], class_name: str, field: str) -> float | None:
    values = [
        float(row[field])
        for row in rows
        if row.get("damage_type") == class_name
        and row.get(field) is not None
    ]
    return round(sum(values), 10) if values else None


def write_csv(
    path: str | Path,
    rows: list[dict[str, Any]],
    fieldnames: Sequence[str] | None = None,
) -> None:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    columns = list(fieldnames) if fieldnames is not None else fieldnames_from_rows(rows)
    with path.open("w", encoding="utf-8-sig", newline="") as stream:
        if not columns:
            return
        writer = csv.DictWriter(stream, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def fieldnames_from_rows(rows: Sequence[Mapping[str, Any]]) -> list[str]:
    columns: list[str] = []
    for row in rows:
        for key in row:
            if key not in columns:
                columns.append(str(key))
    return columns


def json_ready(value: Any) -> Any:
    if isinstance(value, Mapping):
        return {str(key): json_ready(item) for key, item in value.items()}
    if isinstance(value, (list, tuple)):
        return [json_ready(item) for item in value]
    if isinstance(value, np.ndarray):
        return value.tolist()
    if isinstance(value, (np.bool_, bool)):
        return bool(value)
    if isinstance(value, np.integer):
        return int(value)
    if isinstance(value, np.floating):
        return float(value)
    if isinstance(value, Path):
        return str(value)
    return value
