from __future__ import annotations

import csv
import json
from pathlib import Path

import numpy as np
from openpyxl import load_workbook

from uav_rgb.config import load_config
from uav_rgb.reporting import (
    DETAIL_COLUMNS,
    PRIMARY_COLUMNS,
    save_image_result,
    write_combined_outputs,
)


PROJECT_ROOT = Path(__file__).resolve().parents[2]
EXPECTED_OUTPUT_COLUMNS = [
    "image",
    "damage_id",
    "damage_type",
    "damage_name_ko",
    "pixel_nodes_json",
    "world_center_x_m",
    "world_center_y_m",
    "world_center_z_m",
    "length_px",
    "length_m",
    "width_px",
    "width_m",
    "area_m2",
]


def test_output_schema_is_locked_to_the_requested_columns():
    schema = json.loads(
        (PROJECT_ROOT / "03_Processing/schemas/damage_results.schema.json").read_text(
            encoding="utf-8"
        )
    )
    assert schema["additionalProperties"] is False
    assert schema["required"] == EXPECTED_OUTPUT_COLUMNS
    assert list(schema["properties"]) == EXPECTED_OUTPUT_COLUMNS


class FakeGeoContext:
    def image_regions_to_world3d(
        self, rows, pred_mask=None, *, class_masks=None, instance_label_maps=None
    ):
        results = []
        for index, row in enumerate(rows):
            crack = int(row["class_id"]) == 1
            results.append(
                {
                    "world_x_m": 243000.0 + index,
                    "world_y_m": 431000.0 + index,
                    "world_z_m": 100.0,
                    "xyz_valid": True,
                    "nodes_image_xy_json": "[[1.0,2.0],[3.0,4.0]]",
                    "nodes_world_xyz_json": "[[243000,431000,100]]",
                    "node_xyz_hit_ratio": 1.0,
                    "length_m": 0.12 if crack else None,
                    "width_m": 0.004 if crack else None,
                    "area_m2": None if crack else 0.02,
                    "gsd_length_m_per_px": 0.001,
                    "gsd_width_m_per_px": 0.001,
                    "gsd_area_m2_per_px": 0.000001,
                    "measurement_valid": True,
                    "measurement_quality": "complete",
                    "measurement_is_approximate": True,
                    "measurement_miss_reason": None,
                }
            )
        return results

    def to_dict(self):
        return {"method": "test_mesh_ray", "surface_measurement": {"approximate": True}}


def test_pdf_style_columns_and_excel_are_generated(tmp_path):
    assert PRIMARY_COLUMNS == EXPECTED_OUTPUT_COLUMNS
    assert DETAIL_COLUMNS == EXPECTED_OUTPUT_COLUMNS
    config = load_config(
        PROJECT_ROOT / "03_Processing/configs/daechung_aug512.yaml"
    )
    image_path = tmp_path / "sample.jpg"
    image_path.write_bytes(b"placeholder; exporter only records this path")
    masks = {
        "CRC": np.pad(np.ones((2, 5), dtype=bool), ((2, 8), (2, 5))),
        "DLM": np.pad(np.ones((3, 3), dtype=bool), ((7, 2), (6, 3))),
        "SPL": np.zeros((12, 12), dtype=bool),
    }
    prediction = {
        "masks": masks,
        "image_width": 12,
        "image_height": 12,
        "checkpoint": {"sha256": config.expected_checkpoint_sha256},
        "inference": {"edge_handling": {"source_resized": False}},
        "classes": {},
    }
    result = save_image_result(
        image_rgb=np.zeros((12, 12, 3), dtype=np.uint8),
        image_path=image_path,
        prediction=prediction,
        config=config,
        geo3d_context=FakeGeoContext(),
        output_dir=tmp_path / "run",
        min_instance_area_px=1,
    )
    assert [row["damage_name_ko"] for row in result["rows"]] == ["균열", "박리"]
    assert all(list(row) == PRIMARY_COLUMNS for row in result["rows"])
    assert result["rows"][0]["length_m"] == 0.12
    assert result["rows"][1]["area_m2"] == 0.02
    with Path(result["paths"]["details_csv"]).open(
        encoding="utf-8-sig", newline=""
    ) as stream:
        assert next(csv.reader(stream)) == EXPECTED_OUTPUT_COLUMNS
    assert "pixel_summary_csv" not in result["paths"]
    assert not list((tmp_path / "run/tables/per_image").glob("*_pixel_summary.csv"))

    combined = write_combined_outputs(
        rows=result["rows"],
        run_metadata={"model": "aug512", "source_resized": False},
        output_dir=tmp_path / "run",
    )
    workbook = load_workbook(combined["excel"], read_only=True)
    assert workbook.sheetnames == ["Damage_Details"]
    headers = [cell.value for cell in next(workbook["Damage_Details"].iter_rows())]
    assert headers == EXPECTED_OUTPUT_COLUMNS
    assert workbook["Damage_Details"]["D2"].value == "균열"
    assert "image_summary_csv" not in combined
    assert not (tmp_path / "run/tables/image_summary.csv").exists()
